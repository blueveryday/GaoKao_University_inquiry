# 批量处理../src/special中的json专业数据文件
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Border, Font
from openpyxl.worksheet.hyperlink import Hyperlink

def read_json_file(file_path):
    """尝试以不同的编码读取JSON文件"""
    encodings = ['utf-8', 'utf-8-sig', 'gbk']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return json.load(f)
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    raise ValueError(f"无法解码文件: {file_path}")

def save_to_excel(base_folder, output_excel_file):
    """遍历文件夹中的所有JSON文件，并将数据保存为一个Excel文件"""
    all_data = []

    # 遍历所有子文件夹
    for root, dirs, files in os.walk(base_folder):
        for file in files:
            if file.endswith('.json'):
                json_file_path = os.path.join(root, file)
                
                # 读取JSON文件并提取数据
                try:
                    data = read_json_file(json_file_path)
                except ValueError as e:
                    print(e)
                    continue
                
                items = data.get('data', {}).get('item', [])

                # 遍历每个项目并汇总到列表
                for item in items:
                    level1_name = item.get('level1_name', '未知专业层次')
                    level2_name = item.get('level2_name', '未知专业门类')
                    level3_name = item.get('level3_name', '未知专业大类')
                    name = item.get('name', '未知专业')
                    spcode = str(item.get('spcode', '未知专业代码'))  # 确保转换为字符串
                    limit_year = item.get('limit_year', '未知修业年限')
                    degree = item.get('degree', '未知授予学位')
                    boy_rate = item.get('boy_rate', '0')
                    girl_rate = item.get('girl_rate', '0')
                    fivesalaryavg = item.get('fivesalaryavg', '未知平均薪酬')
                    salaryavg = item.get('salaryavg', '未知平均年薪')
                    special_id = item.get('special_id', '未知')
                    school_url = f"https://www.gaokao.cn/special/{special_id}?special_type=3"

                    # 确保将所有数值类型转换为字符串，然后再进行检查
                    if isinstance(fivesalaryavg, (int, float)):
                        fivesalaryavg = str(fivesalaryavg)
                    if isinstance(salaryavg, (int, float)):
                        salaryavg = str(salaryavg)

                    # 计算男女比例
                    gender_ratio = f'"{boy_rate}:{girl_rate}"'

                    # 将数据添加到列表
                    all_data.append([
                        level1_name,
                        level2_name,
                        level3_name,
                        name,
                        spcode,
                        limit_year,
                        degree,
                        gender_ratio,  # 直接写入男女比例，用双引号括起来
                        fivesalaryavg if fivesalaryavg.replace('.', '', 1).isdigit() else None,
                        salaryavg if salaryavg.replace('.', '', 1).isdigit() else None,
                        school_url
                    ])

    # 创建目标文件夹路径
    output_folder = os.path.dirname(output_excel_file)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 创建一个DataFrame并写入Excel文件
    df = pd.DataFrame(all_data, columns=[
        "专业层次", "专业门类", "专业大类", "专业名称", "专业代码", "修业年限", 
        "授予学位", "男女比例", "平均薪酬", "平均年薪", "开设院校"
    ])
    
    # 将"平均薪酬"和"平均年薪"列转换为数值格式，非数字值将被设置为NaN
    df['平均薪酬'] = pd.to_numeric(df['平均薪酬'], errors='coerce')
    df['平均年薪'] = pd.to_numeric(df['平均年薪'], errors='coerce')
    
    # 将"专业代码"列转换为文本格式
    df['专业代码'] = df['专业代码'].astype(str)

    # 保存到Excel文件
    df.to_excel(output_excel_file, index=False, engine='openpyxl')

    # 使用openpyxl打开刚刚保存的Excel文件
    wb = load_workbook(output_excel_file)
    ws = wb.active

    # 定义对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')

    # 定义千位分隔样式
    number_format_style = NamedStyle(name='number_format', number_format='#,##0')

    # 定义无边框样式
    no_border_style = NamedStyle(name='no_border', border=Border())

    # 定义加粗字体样式
    bold_font = Font(bold=True)

    # 对齐表头
    header_titles = {
        'A': '专业层次',
        'B': '专业门类',
        'E': '专业代码',
        'F': '修业年限',
        'H': '男女比例',
        'I': '平均薪酬',
        'J': '平均年薪',
        'K': '开设院校'
    }

    for col_letter, header_name in header_titles.items():
        col_index = df.columns.get_loc(header_name) + 1  # 获取列索引，从1开始
        cell = ws.cell(row=1, column=col_index)
        cell.font = bold_font
        if col_letter in ['I', 'J', 'K']:
            cell.alignment = right_alignment
        else:
            cell.alignment = center_alignment

    # 对数据列对齐
    for col in ['I', 'J', 'K']:
        for cell in ws[col]:
            cell.alignment = right_alignment
            cell.style = number_format_style

    # 对"专业代码"列设置文本格式
    for cell in ws['E']:
        cell.number_format = '@'
        cell.alignment = center_alignment

    # 对"A", "B", "E", "F", "H"列设置居中对齐
    for col in ['A', 'B', 'E', 'F', 'H']:
        for cell in ws[col]:
            cell.alignment = center_alignment

    # 对"K"列设置右对齐
    for cell in ws['K']:
        cell.alignment = right_alignment
        cell.style = number_format_style

    # 将"开设院校"列的数据设置为链接
    for row in ws.iter_rows(min_row=2, max_col=11, max_row=ws.max_row):
        url_cell = row[10]  # "开设院校"列在第11列（从0开始计数）
        if url_cell.value and isinstance(url_cell.value, str) and url_cell.value.startswith('http'):
            url_cell.hyperlink = Hyperlink(ref=url_cell.value, target=url_cell.value)
            url_cell.style = no_border_style

    # 移除标题行的边框
    for cell in ws[1]:
        cell.border = Border()  # 移除边框

    # 重新应用加粗字体样式到指定列的标题
    for col_letter in ['I', 'J', 'K']:
        cell = ws[f'{col_letter}1']
        cell.font = bold_font

    # 保存Excel文件
    wb.save(output_excel_file)
    print(f"\n专业数据已成功保存到 {output_excel_file}\n")
    input("按 Enter 退出")

# 设置基本路径为../src/special，Excel文件保存到../csv/专业数据/专业数据.xlsx
base_folder = '../src/special_info'
output_excel_file = '../csv/专业数据/2024年专业数据汇总.xlsx'
save_to_excel(base_folder, output_excel_file)
