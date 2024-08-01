import os
import json
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.worksheet.hyperlink import Hyperlink

def read_json_file(file_path):
    """尝试以不同的编码读取JSON文件"""
    encodings = ['utf-8', 'utf-8-sig', 'gbk']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return json.load(f)
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    raise ValueError(f"无法解码文件: {file_path}")

def clean_text(text):
    """清理文本中的HTML标签和多余的空格"""
    if isinstance(text, str):  # 确保 text 是字符串类型
        # 去掉 HTML 标签
        text = re.sub(r'<[^>]+>', '', text)
        # 去掉多余的空格和换行符
        text = ' '.join(text.split())
    return text

def process_json_file(json_file_path):
    """处理单个JSON文件，提取数据"""
    try:
        data = read_json_file(json_file_path)
    except ValueError as e:
        print(e)
        return None
    
    item = data.get('data', {})

    # 处理 'jobrate'
    jobrate_data = item.get('jobrate', [])
    jobrate = []
    for entry in jobrate_data:
        year = clean_text(entry.get('year', '未知年份'))
        rate = clean_text(entry.get('rate', '未知就业率'))
        jobrate.append(f"{year}年: {rate}")
    jobrate_text = '\n'.join(jobrate)  # 用换行符连接所有条目

    # 处理 'jobdetail'
    jobdetail = item.get('jobdetail', {})
    job_distribution = []
    area_distribution = []
    position_distribution = []

    # 就业行业分布
    jobdetail_1 = jobdetail.get('1', [])
    for entry in jobdetail_1:
        name = clean_text(entry.get('name', '未知行业'))
        rate = clean_text(entry.get('rate', '未知占比'))
        job_distribution.append(f"{name}: {rate}%")
    job_distribution_text = '\n'.join(job_distribution)

    # 就业地区分布
    jobdetail_2 = jobdetail.get('2', [])
    for entry in jobdetail_2:
        area = clean_text(entry.get('area', '未知地区'))
        rate = clean_text(entry.get('rate', '未知占比'))
        area_distribution.append(f"{area}: {rate}%")
    area_distribution_text = '\n'.join(area_distribution)

    # 就业岗位分布及具体职位
    jobdetail_3 = jobdetail.get('3', [])
    for entry in jobdetail_3:
        detail_pos = clean_text(entry.get('detail_pos', '未知岗位'))
        rate = clean_text(entry.get('rate', '未知占比'))
        detail_job = clean_text(entry.get('detail_job', '未知职位'))
        name = clean_text(entry.get('name', '未知行业'))
        position_distribution.append(f"{detail_pos}: {rate}%\n具体职位：{detail_job}\n所在行业：{name}\n")
    position_distribution_text = '\n'.join(position_distribution)

    # 获取 'professionalsalary' 中的 'salaryavg'
    salary_data = item.get('professionalsalary', {})   
    return [
        clean_text(item.get('level1_name', '未知专业层次')),
        clean_text(item.get('type', '未知专业门类')),
        clean_text(item.get('type_detail', '未知专业大类')),
        clean_text(item.get('name', '未知专业')),
        clean_text(item.get('code', '未知专业代码')),
        clean_text(item.get('limit_year', '未知修业年限')),
        clean_text(item.get('degree', '未知授予学位')),
        clean_text(salary_data.get('salaryavg', '未知毕业五年月薪')),
        clean_text(item.get('salaryavg', '未知平均年薪')),
        clean_text(item.get('rate', '未知男女比例')),
        clean_text(item.get('rate2', '未知文理比例')),
        clean_text(item.get('sel_adv', '未知选科建议')),
        clean_text(item.get('direction', '未知考研方向')),
        clean_text(item.get('celebrity', '未知社会名人')),
        clean_text(item.get('course', '未知专业课程')),
        clean_text(item.get('is_what', '未知专业介绍')),
        clean_text(item.get('learn_what', '未知学习内容')),
        clean_text(item.get('do_what', '未知未来就业')),
        clean_text(item.get('content', '未知专业详解')),
        clean_text(item.get('job', '未知专业就业方向')),
        jobrate_text,  # 已格式化的 'jobrate'
        job_distribution_text,  # 就业行业分布
        area_distribution_text,  # 就业地区分布
        position_distribution_text,  # 就业岗位分布
        #f"https://static-data.gaokao.cn/www/2.0/special/{clean_text(item.get('id', '未知'))}/pc_special_detail.json", 专业json文件地址。
        f"https://www.gaokao.cn/special/{clean_text(item.get('id', '未知'))}?special_type=1",
        f"https://www.gaokao.cn/special/{clean_text(item.get('id', '未知'))}?special_type=3"
    ]

def save_to_excel(base_folder, output_excel_file):
    """遍历文件夹中的所有JSON文件，并将数据保存为一个Excel文件"""
    all_data = []

    # 遍历所有子文件夹
    for root, dirs, files in os.walk(base_folder):
        for file in files:
            if file.endswith('.json'):
                json_file_path = os.path.join(root, file)
                
                # 处理JSON文件并提取数据
                row_data = process_json_file(json_file_path)
                if row_data:
                    all_data.append(row_data)
    
    # 创建目标文件夹路径
    output_folder = os.path.dirname(output_excel_file)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 创建一个DataFrame
    df = pd.DataFrame(all_data, columns=[
        "专业层次", "专业门类", "专业大类", "专业名称", "专业代码", "修业年限", "授予学位", "毕业五年月薪", "平均年薪", "男女比例", "文理比例", "选科建议", "考研方向", "社会名人", 
        "专业课程", "专业介绍", "学习内容", "未来就业", "专业详解", "专业就业方向", "近三年就业率", "就业行业分布", "就业地区分布", "就业岗位分布", "专业详解网址", "开设院校网址"
    ])
    
    # 将DataFrame写入Excel文件
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 设置列宽
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # 设置表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置单元格对齐方式
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in worksheet.iter_rows(min_row=2, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = alignment

        # 设置边框
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        # 设置千位分隔格式
        for col in ['H', 'I']:  # H列是“毕业五年月薪”，I列是“平均年薪”
            for cell in worksheet[col]:
                if cell.row > 1:  # 跳过表头
                    cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
        # 设置超链接
        for row in worksheet.iter_rows(min_row=2, max_col=len(df.columns)):
            for cell in row:
                if cell.column_letter in ['Y', 'Z']:  # Y列是“专业详解网址”，Z列是“开设院校网址”
                    if cell.value:
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=cell.value)
                        cell.style = "Hyperlink"
                        
    print(f"\n专业数据已成功保存到 {output_excel_file}\n")
    input("按 Enter 键退出...")

base_folder = '../src/special/summary'
output_excel_file = '../csv/专业数据/2024年_Major_Summary.xlsx'
save_to_excel(base_folder, output_excel_file)
