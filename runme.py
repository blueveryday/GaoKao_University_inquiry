# 编辑于2024-07-25 23:45

import json
import requests
import csv
import os
import platform
import sys
import pandas as pd
import plotly.graph_objs as go
import plotly.express as px
import plotly.io as pio
import colorama
from colorama import Fore, Style, init
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

def download_file(url, local_filename):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    with requests.get(url, stream=True, headers=headers) as response:
        response.raise_for_status()
        with open(local_filename, 'wb') as f:
            for chunk in response.iter_content(chunk_size=819200):
                if chunk:
                    f.write(chunk)    

def search_school_id(csv_file, keyword):
    try:
        # 读取CSV文件
        df = pd.read_csv(csv_file, encoding='utf-8')
        # 进行模糊查询
        result = df[df['学校名称'].str.contains(keyword, case=False)]
        # 重置索引并删除原索引列
        result = result.reset_index(drop=True)
        # 打印查询结果
        if not result.empty:
            print(Fore.RED + "\n查询结果：" + Style.RESET_ALL)  # 将查询结果文字颜色设置为红色
            # 打印学校名称和学校ID号的标题
            print(Fore.GREEN + "学校名称".ljust(30) + "学校ID号" + Style.RESET_ALL)
            # 打印查询结果
            if not result.empty:
                for idx, row in result.iterrows():
                    print(f"{idx:<5}", end="")  # 打印重置后的索引号
                    print(Fore.GREEN + f"{row['学校名称']:<30}" + Fore.RESET + f"{row['学校ID号']:<10}\n")  # 将学校名称左对齐，学校ID号左对齐，并将学校名称和学校ID号颜色设置为绿色
        else:
            input(Fore.RED + "未找到包含关键字的学校名称，按  Enter 键返回...\n" + Style.RESET_ALL)
            search_menu()  # 返回子菜单
    except Exception as e:
        print("程序出现异常：", e)

def search_province_code(csv_file, keyword):
    # 通过省市区关键字查询省市区数字代码

    try:
        # 读取CSV文件
        df = pd.read_csv(csv_file, encoding='utf-8')
        # 进行模糊查询
        result = df[df.iloc[:, 0].str.contains(keyword, case=False)]
        # 重置索引并删除原索引列
        result = result.reset_index(drop=True)
        # 打印查询结果
        if not result.empty:
            print(Fore.RED + "\n查询结果：" + Style.RESET_ALL)  # 将查询结果文字颜色设置为红色
            for idx, row in result.iterrows():
                print(f"{idx:<5}", end="")  # 打印重置后的索引号
                print(Fore.GREEN + f"{row.iloc[0]:<30}" + Fore.RESET + f"{row.iloc[1]:<10}\n" + Style.RESET_ALL)  # 将省市区名称左对齐，省市区代码左对齐，并将省市区名称和省市区代码颜色设置为绿色
        else:
            input(Fore.RED + "未找到包含关键字的省市区名称，按  Enter 键返回...\n" + Style.RESET_ALL)
            search_menu()  # 返回子菜单
    except Exception as e:
        print("程序出现异常：", e)

def search_json_data(filepath, score):
    while True:
        # 读取JSON文件并搜索数据
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            # 检查输入的高考分数是否在JSON数据的search字段中作为键存在
            if score in data["data"]["search"]:
                search_results = [data["data"]["search"][score]]
                break  # 如果找到结果，则退出循环
            else:
                print(Fore.RED + "你输入的高考分数有误，请重新输入。" + Style.RESET_ALL)
                score = input(Fore.GREEN + " ※ 请重新输入查询的高考分数：" + Style.RESET_ALL)
    # 返回搜索结果
    return search_results

def generate_score_ranking_table(filepath, local_type_id, province_name, local_province_id, year):
    # 读取 JSON 数据
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
        scores = [item['score'] for item in data["data"]["list"]]
        nums = [item['num'] for item in data["data"]["list"]]
        totals = [item['total'] for item in data["data"]["list"]]  # 添加总数数据
        appositive_fractions = [item['appositive_fraction'] for item in data["data"]["list"]]  # 获取历史同位次考生得分数据
        rank_ranges = [item['rank_range'] for item in data["data"]["list"]]  # 获取排名区间数据

    # 创建工作簿
    wb = Workbook()
    # 删除默认的空白工作表
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 添加趋势图工作表并设置内容
    #ws1 = wb.active
    #ws1.title = "趋势图"
    #trend_data = [
    #    ["年份", "分数", "名次MAX", "名次MIN", "中位段", "位次段"],
    #    ['=LEFT(一分一段表!J1,5)', '=VLOOKUP(B5, 一分一段表!$B:$K, 9, FALSE)', '=--LEFT(TRIM(CLEAN(F2)), FIND("-", TRIM(CLEAN(F2))) - 1)', '=--RIGHT(TRIM(CLEAN(F2)), LEN(TRIM(CLEAN(F2))) - FIND("-", TRIM(CLEAN(F2))))', '=INT((C2+D2)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 10, FALSE))'],
    #    ['=LEFT(一分一段表!H1,5)', '=VLOOKUP(B5, 一分一段表!$B:$K, 7, FALSE)', '=--LEFT(TRIM(CLEAN(F3)), FIND("-", TRIM(CLEAN(F3))) - 1)', '=--RIGHT(TRIM(CLEAN(F3)), LEN(TRIM(CLEAN(F3))) - FIND("-", TRIM(CLEAN(F3))))', '=INT((C3+D3)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 8, FALSE))'],
    #    ['=LEFT(一分一段表!F1,5)', '=VLOOKUP(B5, 一分一段表!$B:$K, 5, FALSE)', '=--LEFT(TRIM(CLEAN(F4)), FIND("-", TRIM(CLEAN(F4))) - 1)', '=--RIGHT(TRIM(CLEAN(F4)), LEN(TRIM(CLEAN(F4))) - FIND("-", TRIM(CLEAN(F4))))', '=INT((C4+D4)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 6, FALSE))'],
    #    ['=一分一段表!A1', '=--630', '=--LEFT(TRIM(CLEAN(F5)), FIND("-", TRIM(CLEAN(F5))) - 1)', '=--RIGHT(TRIM(CLEAN(F5)), LEN(TRIM(CLEAN(F5))) - FIND("-", TRIM(CLEAN(F5))))', '=INT((C5+D5)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 3, FALSE))']
    #]
    #for row in trend_data:
    #    ws1.append(row)
    #for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    #    for cell in row:
    #        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 添加数据筛选工作表并设置内容
    ws2 = wb.create_sheet(title="数据筛选")
    filter_data = [
        ["年份", "分数", "名次MAX", "名次MIN", "中位段", "位次段"],
        ['=LEFT(一分一段表!J1,5)', '=VLOOKUP(B5, 一分一段表!$B:$K, 9, FALSE)', '=--LEFT(TRIM(CLEAN(F2)), FIND("-", TRIM(CLEAN(F2))) - 1)', '=--RIGHT(TRIM(CLEAN(F2)), LEN(TRIM(CLEAN(F2))) - FIND("-", TRIM(CLEAN(F2))))', '=INT((C2+D2)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 10, FALSE))'],
        ['=LEFT(一分一段表!H1,5)', '=VLOOKUP(B5, 一分一段表!$B:$K, 7, FALSE)', '=--LEFT(TRIM(CLEAN(F3)), FIND("-", TRIM(CLEAN(F3))) - 1)', '=--RIGHT(TRIM(CLEAN(F3)), LEN(TRIM(CLEAN(F3))) - FIND("-", TRIM(CLEAN(F3))))', '=INT((C3+D3)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 8, FALSE))'],
        ['=LEFT(一分一段表!F1,5)', '=VLOOKUP(B5, 一分一段表!$B:$K, 5, FALSE)', '=--LEFT(TRIM(CLEAN(F4)), FIND("-", TRIM(CLEAN(F4))) - 1)', '=--RIGHT(TRIM(CLEAN(F4)), LEN(TRIM(CLEAN(F4))) - FIND("-", TRIM(CLEAN(F4))))', '=INT((C4+D4)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 6, FALSE))'],
        ['=一分一段表!A1', '=--630', '=--LEFT(TRIM(CLEAN(F5)), FIND("-", TRIM(CLEAN(F5))) - 1)', '=--RIGHT(TRIM(CLEAN(F5)), LEN(TRIM(CLEAN(F5))) - FIND("-", TRIM(CLEAN(F5))))', '=INT((C5+D5)/2)', '=TRIM(VLOOKUP(B5, 一分一段表!$B:$K, 3, FALSE))']
    ]
    for row in filter_data:
        ws2.append(row)
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 将特定单元格设为红色
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws2['B5'].fill = red_fill
    # 添加一分一段表工作表并设置内容
    ws3 = wb.create_sheet(title="一分一段表")
    title_row = ['分数', '同分人数', '排名区间', '累计人数']
    ls_years = set()
    for item in data["data"]["search"].values():
        for fraction in item["appositive_fraction"]:
            ls_years.add(fraction["year"])
    for ls_year in sorted(ls_years, reverse=True):
        title_row.extend([f'{ls_year}年同位次分数', f'{ls_year}年排名区间'])
    ws3.append([''] + title_row)
    for score, num, total, app_fraction, rank_range in zip(scores, nums, totals, appositive_fractions, rank_ranges):
        row_data = [score, num, rank_range, total]
        for ls_year in sorted(ls_years, reverse=True):
            for fraction in app_fraction:
                if fraction["year"] == ls_year:
                    row_data.extend([fraction["score"], fraction["rank_range"]])
                    break
        row_data = [value if isinstance(value, (int, float)) else float(value) if value.replace('.', '', 1).isdigit() else value for value in row_data]
        ws3.append([''] + row_data)
    ws3['A1'] = f"{year}年"
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存工作簿
    csv_folder = os.path.join("csv", str(province_name))
    os.makedirs(csv_folder, exist_ok=True)
    
    # 映射 local_type_id 到 local_type_display_name 的字典
    local_type_display_name_map = {
        2073: '物理类',
        2074: '历史类',
        1: '理科',
        2: '文科'
    }
    try:
        local_type_id = int(local_type_id)
    except ValueError:
        print(f"Error: local_type_id 的值不正确: {local_type_id}")
        local_type_id = None
    # 获取 local_type_display_name
    if local_type_id is not None:
        local_type_display_name = local_type_display_name_map.get(local_type_id, '未知类型')
    else:
        local_type_display_name = '未知类型'

    yfyd_name = f"{local_type_display_name}"
    filename = f"一分一段表_{yfyd_name}_{province_name}_{year}.xlsx"
    filepath = os.path.join(csv_folder, filename)
    wb.save(filepath)
    # 读取并处理已保存的工作簿
    wb = load_workbook(filepath)
    ws = wb["一分一段表"]
    # 复制第二行的数据
    second_row = [cell.value for cell in ws[2]]
    ws.insert_rows(3)
    for col_num, value in enumerate(second_row, start=1):
        ws.cell(row=3, column=col_num, value=value)
    # 修改 B3 单元格的内容
    original_b3_value = ws["B3"].value
    if original_b3_value and isinstance(original_b3_value, str):
        extracted_value = original_b3_value[:3]
        try:
            integer_value = int(extracted_value)
            ws["B3"] = integer_value
        except ValueError:
            print(f"无法将提取的字符串 '{extracted_value}' 转换为整数。")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(filepath)
    print(f"文件已保存到: {filepath} ")

def get_province_name(local_province_id):
    # 通过省市区代码查询对应的省市区名称
    src_province_file_path = "src/province_id.csv"
    if os.path.exists(src_province_file_path):
        with open(src_province_file_path, 'r', encoding='utf-8-sig') as src_province_file:
            reader = csv.reader(src_province_file)
            for row in reader:
                if row[1] == local_province_id:
                    province_name = row[0]
                    return province_name
    return None

def download_json(year, local_province_id, local_type_id):
    #https://static-data.gaokao.cn/www/2.0/section2021/2024/50/2073/3/lists.json
    url = f"https://static-data.gaokao.cn/www/2.0/section2021/{year}/{local_province_id}/{local_type_id}/3/lists.json"
    response = requests.get(url)
    if response.status_code == 200:
        filename = f"lists_{local_province_id}_{year}_{local_type_id}.json"
        folder_path = os.path.join("src", str(local_province_id), "score_ranking")
        os.makedirs(folder_path, exist_ok=True)
        filepath = os.path.join(folder_path, filename)
        with open(filepath, 'wb') as f:
            f.write(response.content)
        print(f"JSON 文件已下载至 {filepath}")
    else:
        print("\nJSON文件下载失败，请访问https://www.gaokao.cn/colleges/bypart 查看各省分有效的查询年份，并重新输入年份下载，或者检查你的网络连接是否正常。\n")

def search_menu():
    while True:
        print("=========================================")
        print("请选择查询类型：")
        print(Fore.GREEN + " [1] 省市区代码查询" + Style.RESET_ALL)
        print(Fore.GREEN + " [2] 学校ID号查询\n" + Style.RESET_ALL)
        print(Fore.RED + " [0] 返回上一级主菜单\n" + Style.RESET_ALL)
        choice = int(input("请输入选择: "))
        if choice == 1:
            csv_file = "src/province_id.csv"
            keyword = input(Fore.RED + "请输入省市区名称关键字：" + Style.RESET_ALL)
            search_province_code(csv_file, keyword)
        elif choice == 2:
            csv_file = "src/school_id.csv"
            keyword = input(Fore.RED + "请输入学校名称关键字：" + Style.RESET_ALL)
            search_school_id(csv_file, keyword)
        elif choice == 0:
            break  # 退出循环，返回上一级主界面
        else:
            print(Fore.RED + "输入错误，请重新输入正确选项！" + Style.RESET_ALL)
            continue

def score_ranking_menu():
    while True:
        # 初始化 colorama 模块
        init(autoreset=True)
        while True:
            print("=========================================")
            print("一分一段查询（同分人数、排名区间等）：\n")
            print(Fore.GREEN + " [1] 通过高考分数查询（一分一段的同分人数、排名区间、累计人数、历史同位次考生得分）")
            print(Fore.GREEN + " [2] 下载2016 - 2024年度的物理类（理科）、历史类（文科）一分一段JSON数据文件")
            print(Fore.GREEN + " [3] 生成一分一段EXCEL文件")
            print(Fore.GREEN + " [4] 生成一分一段折线图")
            print(Fore.GREEN + " [5] 打开一分一段EXCEL文件\n" + Style.RESET_ALL)
            print(Fore.RED + " [0] 返回上级菜单" + Style.RESET_ALL)
            choice = input("\n请输入选项：")
            if choice == '1':
                global local_province_id, local_type_id, year
                folder_path = os.path.join("src", str(local_province_id), "score_ranking")
                os.makedirs(folder_path, exist_ok=True)
                filepath = os.path.join(folder_path, f"lists_{local_province_id}_{year}_{local_type_id}.json")  # 设置文件路径
                province_name_results = get_province_name(local_province_id)  # 查询省市区名称
                if not os.path.exists(filepath):
                    os.system("cls" if os.name == "nt" else "clear")
                    print(Fore.RED + f"\n你查询的{year}年一分一段JSON文件不存在，请先自定义下载文件后再查询！\n" + Style.RESET_ALL)
                    break  # 文件不存在，重新输入
                while True:
                    # 提示输入查询高考分数
                    while True:
                        score_or_rank_input = input(Fore.GREEN + "\n ※ 请输入查询的高考分数：" + Style.RESET_ALL)
                        if not score_or_rank_input.isdigit() or int(score_or_rank_input) < 0 or int(score_or_rank_input) > 750:
                            print(Fore.RED + "错误：请输入0 - 750 之间的整数。" + Style.RESET_ALL)
                        else:
                            break  # 如果输入合法，则退出循环
                    search_results = search_json_data(filepath, score_or_rank_input)
                    if search_results:
                        print(Fore.GREEN + f"\n查询的结果如下（省市区：{province_name_results}，年份：{year}）：\n" + Style.RESET_ALL)
                        for result in search_results:
                            print(f"高考分数: {result['score']}")
                            print(f"同分人数: {result['num']}")
                            print(f"排名区间: {result['rank_range']}")
                            print(f"累计人数: {result['total']}")
                            print(f"批次: {result['batch_name']}\n")
                            # 获取appositive_fraction数组中的参数
                            for app_fraction in result["appositive_fraction"]:
                                ls_year1 = app_fraction["year"]
                                ls_score1 = app_fraction["score"]
                                ls_rank_range1 = app_fraction["rank_range"]
                                print(f"{ls_year1}年同位次考生得分:{ls_score1}, 排名区间：{ls_rank_range1}")
                        print()  # 每个结果之间用空行分隔
                    else:
                        print(Fore.RED + "未找到与输入内容匹配的数据。" + Style.RESET_ALL)
                    # 提示询问用户是否继续查询
                    while True:
                        continue_search = input(Fore.GREEN + "是否继续查询？（Y/n，默认按 Enter 键继续）：" + Style.RESET_ALL) or 'y'
                        print()
                        if continue_search.lower() == 'y':
                            break  # 跳出内层循环，继续查询
                        elif continue_search.lower() == 'n':
                            break  # 跳出内层循环，返回上级菜单
                        else:
                            print(Fore.RED + "错误：请输入 Y 或 n 。" + Style.RESET_ALL)
                    if continue_search.lower() == 'n':
                        os.system('cls' if os.name == 'nt' else 'clear')
                        break  # 返回上级菜单
                break  # 跳出当前层循环，返回到主菜单
            elif choice == '2':
                while True:
                    gk_year = input(Fore.GREEN + "\n请输入年份" + Fore.RED + "（2016 - 2028之间的年份，默认值为2024）: " + Style.RESET_ALL) or "2024"
                    if not gk_year:
                        gk_year = "2024"
                    elif not gk_year.isdigit() or int(gk_year) not in range(2016, 2028):    # 设置查询的年份值范围最大值可修改
                        print(Fore.RED + "错误：请输入2016 - 2028之间的有效年份。" + Style.RESET_ALL)
                        continue
                    else:
                        break
                while True:
                    local_province_id = input(Fore.GREEN + " ※ 请输入省市区代码" + Fore.RED + "（例如 50，默认值为50）: " + Style.RESET_ALL) or "50"
                    if get_province_name(local_province_id) is None:
                        print(Fore.RED + "无效的省市区代码，请重新输入。\n" + Style.RESET_ALL)
                        continue
                        os.system('cls' if os.name == 'nt' else 'clear')
                    else:
                        break
                if int(gk_year) >= 2021:
                    while True:
                        local_type_id = input(Fore.GREEN + "请输入物理、历史类代码" + Fore.RED + "（2073 代表物理类，2074 代表历史类，默认值为2073）: " + Style.RESET_ALL) or "2073"
                        if local_type_id not in ['2073', '2074']:
                            print(Fore.RED + "错误：2021以后的年份（含），本地类别代码只能为 2073（物理类）或 2074（历史类）。" + Style.RESET_ALL)
                            continue
                        else:
                            break
                elif int(gk_year) <= 2020:
                    while True:
                        local_type_id = input(Fore.GREEN + "请输入文、理科代码" + Fore.RED + "（1 代表理科，2代表文科），默认值为1: " + Style.RESET_ALL) or "1"
                        if local_type_id not in ['1', '2']:
                            print(Fore.RED + "错误：2021之前的年份（不含），本地类别代码只能为 1（文科）或 2（理科）。" + Style.RESET_ALL)
                            continue
                        else:
                            break
                download_json(gk_year, local_province_id, local_type_id)
                input("按 Enter 键返回。")
                os.system('cls' if os.name == 'nt' else 'clear')
            elif choice == '3':
                while True:
                    year = input(Fore.GREEN + " ※ 请输入年份" + Fore.RED + "（例如 2016 - 2028之间的年份，默认值为2024）: " + Style.RESET_ALL) or "2024"
                    if not year.isdigit() or int(year) not in range(2016, 2028):
                        print(Fore.RED + "错误：请输入2016 - 2028之间的有效年份。" + Style.RESET_ALL)
                        continue
                    else:
                        break  # 如果输入的年份有效，则退出循环
                while True:
                    local_province_id = input(Fore.GREEN + " ※ 请输入省市区代码" + Fore.RED + "（例如 50，默认值为50）: " + Style.RESET_ALL) or "50"
                    if get_province_name(local_province_id) is None:
                        print(Fore.RED + "无效的省市区代码，请重新输入。\n" + Style.RESET_ALL)
                        continue
                        os.system('cls' if os.name == 'nt' else 'clear')
                    else:
                        break
                if int(year) >= 2021:
                    while True:
                        # 提示用户输入并获取 local_type_id
                        local_type_id = input(Fore.GREEN + " ※ 请输入物理、历史类代码" + Fore.RED + "（2021年及之后，2073 代表物理类，2074 代表历史类，默认值为2073）: " + Style.RESET_ALL) or "2073"
                        # 检查 local_type_id 是否在指定的范围内
                        if local_type_id in ["2073", "2074"]:
                            break  # 如果输入正确，跳出循环
                        else:
                            print(Fore.RED + "你输入的数字错误，请按照提示重新输入文理科代码！" + Style.RESET_ALL)
                            print("2021年之后的文理科代码是：" + Fore.RED + "2073 代表物理类，2074 代表历史类。" + Style.RESET_ALL)
                else:
                    while True:
                        # 提示用户输入并获取 local_type_id
                        local_type_id = input(Fore.GREEN + " ※ 请输入文、理科代码" + Fore.RED + "（2021年之前（不含），1 代表理科，2 代表文科，默认值为1）: " + Style.RESET_ALL) or "1"
                        # 检查 local_type_id 是否在指定的范围内

                        if local_type_id in ["1", "2"]:
                            break  # 如果输入正确，跳出循环
                        else:
                            print(Fore.RED + "你输入的数字错误，请按照提示重新输入文理科代码！" + Style.RESET_ALL)
                            print("2021年之前的文理科代码是：" + Fore.RED + "1 代表理科，2 代表文科；" + Style.RESET_ALL)
                # 获取省份名称
                province_name = get_province_name(local_province_id)
                if not province_name:
                    print(Fore.RED + "错误：未找到对应的省份名称。" + Style.RESET_ALL)
                    continue
                folder_path = os.path.join("src", str(local_province_id), "score_ranking")
                os.makedirs(folder_path, exist_ok=True)
                filepath = os.path.join(folder_path, f"lists_{local_province_id}_{year}_{local_type_id}.json")  # 设置文件路径
                # 检查文件是否存在
                if not os.path.exists(filepath):
                    os.system("cls" if os.name == "nt" else "clear")
                    print(Fore.RED + f"\nsrc路径下不存在 {os.path.basename(filepath)} 文件，请重新从选项[2]中下载正确的JSON文件。\n" + Style.RESET_ALL)
                    break  # 返回上级菜单
                # 生成一分一段表并保存为Excel
                generate_score_ranking_table(filepath, local_type_id, province_name, local_province_id, year)
                input(Fore.GREEN + "请按 Enter 键返回上级菜单。" + Style.RESET_ALL)
                os.system("cls" if os.name == "nt" else "clear")  # 清空屏幕命令
            elif choice == '4':
                # 映射 local_type_id 到 local_type_display_name 的字典
                local_type_display_name_map = {
                    2073: '物理类',
                    2074: '历史类',
                    1: '理科',
                    2: '文科'
                }
                try:
                    local_type_id = int(local_type_id)
                except ValueError:
                    print(f"Error: local_type_id 的值不正确: {local_type_id}")
                    local_type_id = None
                # 获取 local_type_display_name
                if local_type_id is not None:
                    local_type_display_name = local_type_display_name_map.get(local_type_id, '未知类型')
                else:
                    local_type_display_name = '未知类型'
                # 读取 Excel 文件
                province_name = get_province_name(local_province_id)
                csv_folder = os.path.join("csv", str(province_name))
                os.makedirs(csv_folder, exist_ok=True)
                filename = f"一分一段表_{local_type_display_name}_{province_name}_{year}.xlsx"
                file_path = os.path.join(csv_folder, filename)
                sheet_name = '一分一段表'
                # 使用 pandas 读取数据
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # 提取数据
                x = df.iloc[1:, 1].tolist()  # 从第二行到最后一行的 B 列数据
                y = df.iloc[1:, 2].tolist()  # 从第二行到最后一行的 C 列数据
                labels = df.iloc[1:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]]  # 从第二行到最后一行的 B、C、D、E、F、G、H、I、J、K 列数据
                # 使用 Plotly 绘制图表
                plt_title_name = f"一分一段折线图 {local_type_display_name} {province_name} {year}"
                fig = go.Figure()
                # 添加折线图
                fig.add_trace(go.Scatter(
                    x=x,
                    y=y,
                    mode='lines+markers',
                    text=labels.apply(lambda row: '<br>'.join([f"{col}: {val}" for col, val in row.items()]), axis=1),
                    hoverinfo='text',
                    line=dict(color='#FF6F00'),
                    marker=dict(color='#FF6F00')
                ))

                # 更新布局
                fig.update_layout(
                    title=plt_title_name,
                    xaxis_title='分数',
                    yaxis_title='同位次人数',
                    xaxis=dict(
                        autorange='reversed',
                        tickvals=[750, 687, 586, 485, 384, 283, 180, 80, 0],
                        showgrid=True,       # 显示横坐标网格线
                        gridcolor='lightgray', # 网格线颜色
                        gridwidth=1          # 网格线宽度
                    ),
                    yaxis=dict(
                        range=[0, max(y) + 100],
                        showgrid=True,       # 显示横坐标网格线
                        gridcolor='lightgray', # 网格线颜色
                        gridwidth=1          # 网格线宽度
                    ),
                    font=dict(family='DengXian', size=14, color='black'),
                    plot_bgcolor='white',
                    hovermode='closest'
                )

                # 更新鼠标悬停标签的样式
                fig.update_traces(
                    hoverlabel=dict(
                        bgcolor='gray',         # 背景颜色
                        font=dict(
                            family='DengXian', # 字体
                            size=12,          # 字体大小
                            color='white',    # 字体颜色
                            weight='bold'     # 字体加粗
                        )
                    )
                )
                # 显示图表
                pio.show(fig)
                os.system('cls' if os.name == 'nt' else 'clear')
                break
            elif choice == '5':
                # 映射 local_type_id 到 local_type_display_name 的字典
                local_type_display_name_map = {
                    2073: '物理类',
                    2074: '历史类',
                    1: '理科',
                    2: '文科'
                }
                try:
                    local_type_id = int(local_type_id)
                except ValueError:
                    print(f"Error: local_type_id 的值不正确: {local_type_id}")
                    local_type_id = None
                # 获取 local_type_display_name
                if local_type_id is not None:
                    local_type_display_name = local_type_display_name_map.get(local_type_id, '未知类型')
                else:
                    local_type_display_name = '未知类型'
                # 获取省份名称
                province_name = get_province_name(local_province_id)
                if not province_name:
                    print(Fore.RED + "错误：未找到对应的省份名称。" + Style.RESET_ALL)
                    continue
                print("正在打开:一分一段表.xlsx文件\n")
                file_path = os.path.join('csv', str(province_name), f"一分一段表_{local_type_display_name}_{province_name}_{year}.xlsx")
                print(f"文件路径: {file_path}\n") 
                if os.path.exists(file_path):
                    try:
                        if os.name == 'nt':  # 如果是Windows系统
                            os.startfile(file_path)
                        elif sys.platform == 'darwin':  # 如果是macOS
                            subprocess.call(['open', file_path])
                        else:  # 如果是Linux系统
                            subprocess.call(['xdg-open', file_path])
                    except Exception as e:
                        print(Fore.RED + f"打开文件时出错: {e}" + Style.RESET_ALL)
                else:
                    print(Fore.RED + "文件不存在！" + Style.RESET_ALL)
                    break
                os.system('cls' if os.name == 'nt' else 'clear')
                break
            elif choice == '0':
                os.system('cls' if os.name == 'nt' else 'clear')
                return
                break  # 返回上一级菜单
            else:
                os.system('cls' if os.name == 'nt' else 'clear')
                print(Fore.RED + "请选择正确的选项。\n" + Style.RESET_ALL)
                continue

def run_code(choice):  
    global local_province_id, local_type_id, school_id, year
    while True:
        if choice == 1:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 院校分数线
            # 读取 school_id.csv 文件获取学校名称
            src_folder = "src"
            src_school_file_name = "school_id.csv"
            src_province_file_name = "province_id.csv"
            src_school_file_path = os.path.join(os.getcwd(), src_folder, src_school_file_name)
            src_province_file_path = os.path.join(os.getcwd(), src_folder, src_province_file_name)
            school_name = "未知学校"  # 默认值，如果找不到对应的学校ID，则使用默认值
            province_name = "未知省份"  # 默认值，如果找不到对应的省市区代码，则使用默认值
            # 查询学校名称
            school_id_name_mapping = {}
            if os.path.exists(src_school_file_path):
                with open(src_school_file_path, 'r', encoding='utf-8-sig') as src_school_file:
                    reader = csv.reader(src_school_file)
                    for row in reader:
                        school_id_name_mapping[row[1]] = row[0]
                        if row[1] == school_id:
                            school_name = row[0]
            # 查询省市区代码对应的省份名称
            province_id_name_mapping = {}
            if os.path.exists(src_province_file_path):
                with open(src_province_file_path, 'r', encoding='utf-8-sig') as src_province_file:
                    reader = csv.reader(src_province_file)
                    for row in reader:
                        province_id_name_mapping[row[1]] = row[0]
                        if row[1] == local_province_id:
                            province_name = row[0]
            # 定义文件夹路径和文件名
            folder_name = "csv"
            download_folder = "download"
            download_folder_path = os.path.join(os.getcwd(), download_folder)
            province_folder = os.path.join(os.getcwd(), folder_name, province_name)
            school_subfolder = os.path.join(province_folder, school_name)
            # 创建文件夹
            os.makedirs(province_folder, exist_ok=True)
            os.makedirs(school_subfolder, exist_ok=True)
            os.makedirs(download_folder_path, exist_ok=True)
            # 定义要下载的文件URL和本地保存路径
            # https://static-data.gaokao.cn/www/2.0/schoolprovincescore/109/2024/50.json
            base_url = 'https://static-data.gaokao.cn/www/2.0/schoolprovincescore'
            url = f"{base_url}/{school_id}/{year}/{local_province_id}.json"
            local_folder = 'download'
            local_filename = os.path.join(local_folder, f"院校分数线_{school_id}_{local_type_id}.json")
            # 创建保存 JSON 文件的文件夹
            if not os.path.exists(local_folder):
                os.makedirs(local_folder)
            # 下载文件
            download_file(url, local_filename)
            # 文件下载完成后，检查文件内容是否包含指定的值
            with open(local_filename, 'r', encoding='utf-8') as f:
                content = json.load(f)
                if content.get("numFound") == 0 or content.get("code") == "0003":
                    print("年份错误，非开启年。请重新输入年份！\n")
                    input("按 Enter 键继续")
                    break
            # 读取 JSON 文件
            items = []
            for type_id, type_data in content['data'].items():
                if 'item' in type_data:
                    items.extend(type_data['item'])
            if not items:
                print("未找到任何数据，请检查下载文件并重试。")
                return
            # 创建保存 CSV 文件的文件夹
            csv_folder = 'csv'
            first_item = items[0]
            school_name = school_id_name_mapping.get(first_item['school_id'], "未知学校")
            province_name = province_id_name_mapping.get(first_item['province_id'], "未知省份")
            local_type_name = first_item['type']
            school_folder_path = os.path.join(csv_folder, province_name, school_name)
            if not os.path.exists(school_folder_path):
                os.makedirs(school_folder_path)
            # 定义CSV文件路径
            csv_file_path = os.path.join(school_folder_path, f"{school_name}_学校代码{school_id}_{province_name}_{year}_院校分数线.csv")
            # 写入CSV文件
            with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
                writer = csv.writer(csv_file)
                # 写入表头
                writer.writerow([
                    "学校名称", "招生年份", "省市区", "文理科类型", "最低分", "最低位次", "录取批次", "招生类型", "省控线"
                ])
                # 定义类型映射
                type_mapping = {
                    "2073": "物理类",
                    "2074": "历史类",
                    "1": "理科",
                    "2": "文科"
                }
                # 提取信息并写入CSV文件
                for item in items:
                    item_province_id = item.get('province_id', '-')
                    province_name = province_id_name_mapping.get(item_province_id, "未知省份")
                    item_type = item.get('type', '-')
                    item_type_name = type_mapping.get(item_type, item_type)
                    writer.writerow([
                        school_name, 
                        item.get('year', '-'), 
                        province_name, 
                        item_type_name, 
                        item.get('min', '-'), 
                        item.get('min_section', '-'), 
                        item.get('local_batch_name', '-'), 
                        item.get('zslx_name', '-'), 
                        item.get('proscore', '-')
                    ])
            print(f"数据已成功保存到 {csv_file_path} 文件中。")
            input("按 Enter 键继续...")
            break
        elif choice == 2:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 专业分数线
            # 读取 school_id.csv 文件获取学校名称
            src_folder = "src"
            src_school_file_name = "school_id.csv"
            src_province_file_name = "province_id.csv"
            src_school_file_path = os.path.join(os.getcwd(), src_folder, src_school_file_name)
            src_province_file_path = os.path.join(os.getcwd(), src_folder, src_province_file_name)
            school_name = "未知学校"  # 默认值，如果找不到对应的学校ID，则使用默认值
            province_name = "未知省份"  # 默认值，如果找不到对应的省市区代码，则使用默认值
            # 查询学校名称
            school_id_name_mapping = {}
            if os.path.exists(src_school_file_path):
                with open(src_school_file_path, 'r', encoding='utf-8-sig') as src_school_file:
                    reader = csv.reader(src_school_file)
                    for row in reader:
                        school_id_name_mapping[row[1]] = row[0]
                        if row[1] == school_id:
                            school_name = row[0]
            # 查询省市区代码对应的省份名称
            province_id_name_mapping = {}
            if os.path.exists(src_province_file_path):
                with open(src_province_file_path, 'r', encoding='utf-8-sig') as src_province_file:
                    reader = csv.reader(src_province_file)
                    for row in reader:
                        province_id_name_mapping[row[1]] = row[0]
                        if row[1] == local_province_id:
                            province_name = row[0]
            # 定义文件夹路径和文件名
            folder_name = "csv"
            download_folder = "download"
            download_folder_path = os.path.join(os.getcwd(), download_folder)
            province_folder = os.path.join(os.getcwd(), folder_name, province_name)
            school_subfolder = os.path.join(province_folder, school_name)
            # 创建文件夹
            os.makedirs(province_folder, exist_ok=True)
            os.makedirs(school_subfolder, exist_ok=True)
            os.makedirs(download_folder_path, exist_ok=True)
            # 定义要下载的文件URL和本地保存路径
            # https://static-data.gaokao.cn/www/2.0/schoolprovincescore/109/2024/50.json
            base_url = 'https://static-data.gaokao.cn/www/2.0/schoolspecialscore'
            url = f"{base_url}/{school_id}/{year}/{local_province_id}.json"
            local_folder = 'download'
            local_filename = os.path.join(local_folder, f"专业分数线_{school_id}_{local_type_id}.json")
            # 创建保存 JSON 文件的文件夹
            if not os.path.exists(local_folder):
                os.makedirs(local_folder)
            # 下载文件
            download_file(url, local_filename)
            # 文件下载完成后，检查文件内容是否包含指定的值
            with open(local_filename, 'r', encoding='utf-8') as f:
                content = json.load(f)
                if content.get("numFound") == 0 or content.get("code") == "0003":
                    print("年份错误，非开启年。请重新输入年份！\n")
                    input("按 Enter 键继续")
                    break
            # 读取 JSON 文件
            items = []
            for type_id, type_data in content['data'].items():
                if 'item' in type_data:
                    items.extend(type_data['item'])
            if not items:
                print("未找到任何数据，请检查下载文件并重试。")
                return
            # 创建保存 CSV 文件的文件夹
            csv_folder = 'csv'
            first_item = items[0]
            school_name = school_id_name_mapping.get(first_item['school_id'], "未知学校")
            province_name = province_id_name_mapping.get(first_item['province'], "未知省份")
            local_type_name = first_item['type']
            school_folder_path = os.path.join(csv_folder, province_name, school_name)
            if not os.path.exists(school_folder_path):
                os.makedirs(school_folder_path)
            # 定义CSV文件路径
            csv_file_path = os.path.join(school_folder_path, f"{school_name}_学校代码{school_id}_{province_name}_{year}_专业分数线.csv")
            # 写入CSV文件
            with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
                writer = csv.writer(csv_file)
                # 写入表头
                writer.writerow([
                    "学校名称", "省市区", "招生年份", "类型", "录取批次", "专业名称", "最低分", "最低位次", "选课要求"
                ])
                # 定义类型映射
                type_mapping = {
                    "2073": "物理类",
                    "2074": "历史类",
                    "1": "理科",
                    "2": "文科"
                }
                # 提取信息并写入CSV文件
                for item in items:
                    item_province_id = item.get('province', '-')
                    province_name = province_id_name_mapping.get(item_province_id, "未知省份")
                    item_type = item.get('type', '-')
                    item_type_name = type_mapping.get(item_type, item_type)
                    writer.writerow([
                        school_name,
                        province_name, 
                        year, 
                        item_type_name, 
                        item.get('local_batch_name', '-'), 
                        item.get('spname', '-'),
                        item.get('min', '-'), 
                        item.get('min_section', '-'),  
                        item.get('sp_info', '-')
                    ])
            print(f"数据已成功保存到 {csv_file_path} 文件中。")
            input("按 Enter 键继续...")
            break
        elif choice == 3:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 招生计划
            # 读取 school_id.csv 文件获取学校名称
            src_folder = "src"
            src_school_file_name = "school_id.csv"
            src_province_file_name = "province_id.csv"
            src_school_file_path = os.path.join(os.getcwd(), src_folder, src_school_file_name)
            src_province_file_path = os.path.join(os.getcwd(), src_folder, src_province_file_name)
            school_name = "未知学校"  # 默认值，如果找不到对应的学校ID，则使用默认值
            province_name = "未知省份"  # 默认值，如果找不到对应的省市区代码，则使用默认值
            # 查询学校名称
            school_id_name_mapping = {}
            if os.path.exists(src_school_file_path):
                with open(src_school_file_path, 'r', encoding='utf-8-sig') as src_school_file:
                    reader = csv.reader(src_school_file)
                    for row in reader:
                        school_id_name_mapping[row[1]] = row[0]
                        if row[1] == school_id:
                            school_name = row[0]
            # 查询省市区代码对应的省份名称
            province_id_name_mapping = {}
            if os.path.exists(src_province_file_path):
                with open(src_province_file_path, 'r', encoding='utf-8-sig') as src_province_file:
                    reader = csv.reader(src_province_file)
                    for row in reader:
                        province_id_name_mapping[row[1]] = row[0]
                        if row[1] == local_province_id:
                            province_name = row[0]
            # 定义文件夹路径和文件名
            folder_name = "csv"
            download_folder = "download"
            download_folder_path = os.path.join(os.getcwd(), download_folder)
            province_folder = os.path.join(os.getcwd(), folder_name, province_name)
            school_subfolder = os.path.join(province_folder, school_name)
            # 创建文件夹
            os.makedirs(province_folder, exist_ok=True)
            os.makedirs(school_subfolder, exist_ok=True)
            os.makedirs(download_folder_path, exist_ok=True)
            # 定义要下载的文件URL和本地保存路径
            # https://static-data.gaokao.cn/www/2.0/schoolspecialplan/109/2024/50.json
            base_url = 'https://static-data.gaokao.cn/www/2.0/schoolspecialplan'
            url = f"{base_url}/{school_id}/{year}/{local_province_id}.json"
            local_folder = 'download'
            local_filename = os.path.join(local_folder, f"招生计划_{school_id}_{local_type_id}.json")
            # 创建保存 JSON 文件的文件夹
            if not os.path.exists(local_folder):
                os.makedirs(local_folder)
            # 下载文件
            download_file(url, local_filename)
            # 文件下载完成后，检查文件内容是否包含指定的值
            with open(local_filename, 'r', encoding='utf-8') as f:
                content = json.load(f)
                if content.get("numFound") == 0 or content.get("code") == "0003":
                    print("年份错误，非开启年。请重新输入年份！\n")
                    input("按 Enter 键继续")
                    break
            # 读取 JSON 文件
            items = []
            for type_id, type_data in content['data'].items():
                if 'item' in type_data:
                    items.extend(type_data['item'])
            if not items:
                print("未找到任何数据，请检查下载文件并重试。")
                return
            # 创建保存 CSV 文件的文件夹
            csv_folder = 'csv'
            first_item = items[0]
            school_name = school_id_name_mapping.get(first_item['school_id'], "未知学校")
            province_name = province_id_name_mapping.get(first_item['province'], "未知省份")
            local_type_name = first_item['type']
            school_folder_path = os.path.join(csv_folder, province_name, school_name)
            if not os.path.exists(school_folder_path):
                os.makedirs(school_folder_path)
            # 定义CSV文件路径
            csv_file_path = os.path.join(school_folder_path, f"{school_name}_学校代码{school_id}_{province_name}_{year}_招生计划.csv")
            # 写入CSV文件
            with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
                writer = csv.writer(csv_file)
                # 写入表头
                writer.writerow([
                    "学校名称", "省市区", "招生年份", "类型", "录取批次", "专业名称", "计划招生", "学制", "学费/年", "选科要求"
                ])
                # 定义类型映射
                type_mapping = {
                    "2073": "物理类",
                    "2074": "历史类",
                    "1": "理科",
                    "2": "文科"
                }
                # 提取信息并写入CSV文件
                for item in items:
                    item_province_id = item.get('province', '-')
                    province_name = province_id_name_mapping.get(item_province_id, "未知省份")
                    item_type = item.get('type', '-')
                    item_type_name = type_mapping.get(item_type, item_type)
                    writer.writerow([
                        school_name,
                        province_name, 
                        year, 
                        item_type_name, 
                        item.get('local_batch_name', '-'), 
                        item.get('spname', '-'),
                        item.get('num', '-'), 
                        item.get('length', '-'),  
                        item.get('tuition', '-'),
                        item.get('sp_info', '-')
                    ])
            print(f"数据已成功保存到 {csv_file_path} 文件中。")
            input("按 Enter 键继续...")
            break
        elif choice == 4:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 查询开设专业
            # 设置请求头中的User-Agent
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
            }
            # 读取 school_id.csv 文件获取学校名称
            src_folder = "src"
            src_school_file_name = "school_id.csv"
            src_province_file_name = "province_id.csv"
            src_school_file_path = os.path.join(
                os.getcwd(), src_folder, src_school_file_name)
            src_province_file_path = os.path.join(
                os.getcwd(), src_folder, src_province_file_name)
            school_name = "未知学校"  # 默认值，如果找不到对应的学校ID，则使用默认值
            province_name = "未知省份"  # 默认值，如果找不到对应的省市区代码，则使用默认值
            # 查询学校名称
            if os.path.exists(src_school_file_path):
               with open(src_school_file_path, 'r', encoding='utf-8-sig') as src_school_file:
                    reader = csv.reader(src_school_file)
                    for row in reader:
                        if row[1] == school_id:
                            school_name = row[0]
                            break
            # 查询省市区代码对应的省份名称
            if os.path.exists(src_province_file_path):
                with open(src_province_file_path, 'r', encoding='utf-8-sig') as src_province_file:
                    reader = csv.reader(src_province_file)
                    for row in reader:
                        if row[1] == local_province_id:
                            province_name = row[0]
                            break
            # 定义文件夹路径和文件名
            folder_name = "csv"
            download_folder = "download"
            download_folder_path = os.path.join(os.getcwd(), download_folder)
            province_folder = os.path.join(os.getcwd(), folder_name, province_name)
            school_subfolder = os.path.join(province_folder, school_name)
            # 创建文件夹
            os.makedirs(province_folder, exist_ok=True)
            os.makedirs(school_subfolder, exist_ok=True)
            os.makedirs(download_folder_path, exist_ok=True)
            # 下载 JSON 文件并保存到 download 文件夹中
            # 地址实例:https://static-data.gaokao.cn/www/2.0/school/109/pc_special.json
            url = f"https://static-data.gaokao.cn/www/2.0/school/{school_id}/pc_special.json"
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                json_file_path = os.path.join(
                    download_folder_path, f"开设专业_{school_id}_pc_special.json")
                with open(json_file_path, 'w', encoding='utf-8') as json_file:
                    json_file.write(response.text)
            else:
                print("JSON 文件下载失败。")
            # 检查请求是否成功
            if response.status_code == 200:
                # 读取 JSON 文件
                data = json.loads(response.text)
                # 提取所需字段并保存为列表
                extracted_data = set()  # 使用集合来存储数据，以去除重复项
                # 提取"data"下的"1"数组里的数据
                for item in data['data'].get('1', []):
                    temp_year = item['year']  # 提取招生年份
                    nation_feature = "国家特色专业" if item.get(
                        'nation_feature') == '1' else ''
                    extracted_data.add((
                        school_name,          # 学校ID
                        item['special_name'],       # 专业名称
                        item['type_name'],          # 层次
                        item['level2_name'],        # 学科门类
                        item['level3_name'],        # 专业类别
                        item['limit_year'],         # 学制
                        item.get('xueke_rank_score', ''),  # 学科等级
                        nation_feature,             # 国家特色专业
                        temp_year                   # 招生年份
                    ))
                # 提取"special_detail"下的"1"数组里的数据
                for item in data['data']['special_detail'].get('1', []):
                    temp_year = item['year']  # 提取招生年份
                    nation_feature = "国家特色专业" if item.get(
                        'nation_feature') == '1' else ''
                    extracted_data.add((
                        school_name,          # 学校ID
                        item['special_name'],       # 专业名称
                        item['type_name'],          # 层次
                        item['level2_name'],        # 学科门类
                        item['level3_name'],        # 专业类别
                        item['limit_year'],         # 学制
                        item.get('xueke_rank_score', ''),  # 学科等级
                        nation_feature,             # 国家特色专业
                        temp_year                   # 招生年份
                    ))
                # 提取"nation_feature"数组里的数据
                for item in data['data']['nation_feature']:
                    temp_year = item['year']  # 提取招生年份
                    nation_feature = "国家特色专业" if item.get(
                        'nation_feature') == '1' else ''
                    extracted_data.add((
                        school_name,                # 学校名称
                        item['special_name'],       # 专业名称
                        item['type_name'],          # 层次
                        item['level2_name'],        # 学科门类
                        item['level3_name'],        # 专业类别
                        item['limit_year'],         # 学制
                        item.get('xueke_rank_score', ''),  # 学科等级
                        nation_feature,             # 国家特色专业
                        temp_year                   # 招生年份
                    ))
                # 在 choice == 4 分支中，使用临时变量存储 year 的值
                temp_year = year
                # 获取招生年份列表中的第一个年份值
                first_year = list(extracted_data)[0][-1]
                # 定义文件名
                file_name = f"{school_name}_学校代码{school_id}_{province_name}{local_province_id}_{temp_year}_开设专业.csv"
                file_path = os.path.join(school_subfolder, file_name)
                # 将数据写入 CSV 文件
                with open(file_path, mode='w', newline='', encoding='utf-8-sig') as file:
                    csv.writer(file).writerows([['学校ID', '专业名称', '层次', '学科门类', '专业类别', '学制', '学科等级', '国家特色专业', '开设专业年份']] + list(extracted_data))
                # print(f"数据已成功保存到 {file_path} 文件中。")
                # #显示文件保存的绝对路径
                print(f"数据已成功保存到 {os.path.relpath(file_path)} 文件中。")  # 显示文件保存的相对路径
            else:
                print("请求失败。")
            input("按 Enter 键继续...") 
            break
        elif choice == 5:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 查询学校学科评估
            # 设置请求头中的User-Agent
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
            }
            # 读取 school_id.csv 文件获取学校名称
            src_folder = "src"
            src_school_file_name = "school_id.csv"
            src_province_file_name = "province_id.csv"
            src_school_file_path = os.path.join(
                os.getcwd(), src_folder, src_school_file_name)
            src_province_file_path = os.path.join(
                os.getcwd(), src_folder, src_province_file_name)
            school_name = "未知学校"  # 默认值，如果找不到对应的学校ID，则使用默认值
            province_name = "未知省份"  # 默认值，如果找不到对应的省市区代码，则使用默认值
            # 查询学校名称
            if os.path.exists(src_school_file_path):
               with open(src_school_file_path, 'r', encoding='utf-8-sig') as src_school_file:
                    reader = csv.reader(src_school_file)
                    for row in reader:
                        if row[1] == school_id:
                            school_name = row[0]
                            break
            # 查询省市区代码对应的省份名称
            if os.path.exists(src_province_file_path):
                with open(src_province_file_path, 'r', encoding='utf-8-sig') as src_province_file:
                    reader = csv.reader(src_province_file)
                    for row in reader:
                        if row[1] == local_province_id:
                            province_name = row[0]
                            break
            # 定义文件夹路径和文件名
            folder_name = "csv"
            download_folder = "download"
            download_folder_path = os.path.join(os.getcwd(), download_folder)
            province_folder = os.path.join(os.getcwd(), folder_name, province_name)
            school_subfolder = os.path.join(province_folder, school_name)
            # 创建文件夹
            os.makedirs(province_folder, exist_ok=True)
            os.makedirs(school_subfolder, exist_ok=True)
            os.makedirs(download_folder_path, exist_ok=True)
            # 下载 JSON 文件并保存到 download 文件夹中
            # 地址实例:https://static-data.gaokao.cn/www/2.0/school/109/xueke_rank.json
            url = f"https://static-data.gaokao.cn/www/2.0/school/{school_id}/xueke_rank.json"
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                json_file_path = os.path.join(
                    download_folder_path, f"学科评估_{school_id}_xueke_rank.json")
                with open(json_file_path, 'w', encoding='utf-8') as json_file:
                    json_file.write(response.text)
            else:
                print("JSON 文件下载失败。")
            # 检查请求是否成功
            if response.status_code == 200:
                # 读取 JSON 文件
                local_folder = 'download'
                local_filename = os.path.join(local_folder, f"学科评估_{school_id}_xueke_rank.json")
                with open(local_filename, encoding='utf-8') as f:
                    data = json.loads(f.read())
                    round_info = data['data']['round']  # 获取 round 信息
                    items = data['data']['item'][0]
                # 提取信息并写入CSV文件
                extracted_data = []
                for item in items:
                    xueke_name = item['xueke_name']  # 学科名称
                    xueke_rank_score = item['xueke_rank_score']  # 评估等级
                    extracted_data.append([xueke_name, xueke_rank_score])
                # 定义文件名
                file_name = f"{school_name}_学校代码{school_id}_{province_name}{local_province_id}_{year}_{'_'.join(round_info)}学科评估.csv"
                file_path = os.path.join(school_subfolder, file_name)
                # 将数据写入 CSV 文件
                with open(file_path, mode='w', newline='', encoding='utf-8-sig') as file:
                    writer = csv.writer(file)
                    writer.writerow(['学科', '评估'])
                    writer.writerows(extracted_data)
                # print(f"数据已成功保存到 {file_path} 文件中。")
                # #显示文件保存的绝对路径
                print(f"数据已成功保存到 {os.path.relpath(file_path)} 文件中。")  # 显示文件保存的相对路径
            else:
                print("请求失败。")
            input("按 Enter 键继续...") 
            break
        elif choice == 6:
            while True:
                # 一键获取学校全部信息，按照指定顺序执行代码
                for code in [1, 2, 3, 4, 5]:
                    try:
                        if code == 1:
                            # 查询院校分数线
                            run_code(code)
                        elif code == 2:
                            # 专业分数线
                            run_code(code)
                        elif code == 3:
                            # 查询招生计划
                            run_code(code)
                        elif code == 4:
                            # 查询开设专业
                            run_code(code)
                        elif code == 5:
                            # 查询学科评估
                            run_code(code)
                        else:
                            print("无效的 code，请检查并重试！")
                    except Exception as e:
                        print(f"执行代码 {code} 时发生错误: {e}")
                # 提示用户是否重新输入学校ID，继续查询
                print("\n查询已完成!")
                while True:
                    continue_search = input(Fore.GREEN + "是否继续一键查询学校全部信息？（Y/n，默认按 Enter 键继续）：" + Style.RESET_ALL) or 'y'
                    if continue_search.lower() == 'y':
                        school_id = input("请输入新的学校ID: ").strip()  # 提示用户输入新的学校ID，继续进行下一个学校的查询。
                        break  # 跳出内层循环，继续查询
                    elif continue_search.lower() == 'n':
                        os.system('cls' if os.name == 'nt' else 'clear')  # 清屏
                        return  # 返回主菜单
                    else:
                        print("无效的输入，请输入 'Y' 或 'n'")        
        elif choice == 7:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 查询省市区代码或学校ID号
            search_menu()
            break
        elif choice == 8:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 重新定义省市区代码、文理科代码、学校ID、录取年份等参数
            while True:
                local_province_id = input(Fore.GREEN + " ※ 请输入省市区代码" + Fore.RED + "（例如 50，默认值为50）: " + Style.RESET_ALL) or "50"
                if get_province_name(local_province_id) is None:
                    print(Fore.RED + "无效的省市区代码，请重新输入。\n" + Style.RESET_ALL)
                    #input("按 Enter 键重新输入...\n")
                    continue
                else:
                    break
            while True:
                year = input(Fore.GREEN + " ※ 请输入年份" + Fore.RED + "（例如 2016 - 2028之间的年份，默认值为2024）: " + Style.RESET_ALL) or "2024"
                if not year.isdigit() or int(year) not in range(2016, 2028):
                    print(Fore.RED + "错误：请输入2016 - 2028之间的有效年份。" + Style.RESET_ALL)
                    continue
                else:
                    break  # 如果输入的年份有效，则退出循环
            if int(year) >= 2021:
                while True:
                    # 提示用户输入并获取 local_type_id
                    local_type_id = input(Fore.GREEN + " ※ 请输入物理、历史类代码" + Fore.RED + "（2021年及之后，2073 代表物理类，2074 代表历史类，默认值为2073）: " + Style.RESET_ALL) or "2073"
                    # 检查 local_type_id 是否在指定的范围内
                    if local_type_id in ["2073", "2074"]:
                        break  # 如果输入正确，跳出循环
                    else:
                        print(Fore.RED + "你输入的数字错误，请按照提示重新输入文理科代码！" + Style.RESET_ALL)
                        print("2021年之后的文理科代码是：" + Fore.RED + "2073 代表物理类，2074 代表历史类。" + Style.RESET_ALL)
            else:
                while True:
                    # 提示用户输入并获取 local_type_id
                    local_type_id = input(Fore.GREEN + " ※ 请输入文、理科代码" + Fore.RED + "（2021年之前（不含），1 代表理科，2 代表文科，默认值为1）: " + Style.RESET_ALL) or "1"
                    # 检查 local_type_id 是否在指定的范围内
                    if local_type_id in ["1", "2"]:
                        break  # 如果输入正确，跳出循环
                    else:
                        print(Fore.RED + "你输入的数字错误，请按照提示重新输入文理科代码！" + Style.RESET_ALL)
                        print("2021年之前的文理科代码是：" + Fore.RED + "1 代表理科，2 代表文科；" + Style.RESET_ALL)
            school_id = input(Fore.GREEN + " ※ 请输入学校ID" + Fore.RED + "(默认：东南大学109)" + Style.RESET_ALL + ":") or "109"
            break
        elif choice == 9:
            os.system("cls" if os.name == "nt" else "clear")  # 清空屏幕命令
            # 查询一分一段
            score_ranking_menu() 
            break
        elif choice == 10:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 清空下载文件夹中的全部文件
            download_folder = 'download'
            if os.path.exists(download_folder):
                for filename in os.listdir(download_folder):
                    file_path = os.path.join(download_folder, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                print("已清空download文件夹\n")
            else:
                print(Fore.RED + "下载文件夹不存在。\n" + Style.RESET_ALL)
            input("按 Enter 键继续...")  
            break
        elif choice == 11:
            os.system('cls' if os.name == 'nt' else 'clear')
            # 更新学校id
            url = 'https://static-data.gaokao.cn/www/2.0/school/school_code.json'
            local_filename = 'download/school_id.json'
            if not os.path.exists('download'):
                os.makedirs('download')
            download_file(url, local_filename)
            with open('download/school_id.json', 'r', encoding='utf-8') as file:
                json_string = file.read()
            parsed_data = json.loads(json_string)
            if not os.path.exists('src'):
                os.makedirs('src')
            with open('src/school_id.csv', mode='w', newline='', encoding='utf-8-sig') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerow(["学校名称", "学校ID号"])
                school_data = parsed_data["data"]
                for key, value in school_data.items():
                    school_id = value["school_id"]
                    name = value["name"]
                    writer.writerow([name, school_id])
            print(f"数据已成功保存到src文件夹中，文件名为:school_id.csv。")
            input("按 Enter 键继续...")
            break
        elif choice == 12:
            def csv_save_as_xlsx(data_path, output_path):
                for dirpath, dirnames, filenames in os.walk(data_path):
                    for fname in filenames:
                        if fname.endswith('.csv'):
                            file_name = os.path.join(dirpath, fname)
                            df = pd.read_csv(file_name)
                            # 构建新的文件夹路径
                            relative_path = os.path.relpath(dirpath, data_path)
                            new_dir = os.path.join(output_path, relative_path)
                            # 确保目标目录存在
                            os.makedirs(new_dir, exist_ok=True)
                            # 构建新文件路径
                            new_file_path = os.path.join(new_dir, f"{os.path.splitext(fname)[0]}.xlsx")
                            df.to_excel(new_file_path, index=False)
                            print(f'{file_name} 转换为 {new_file_path} 成功')
                            print('=========================================')
            # 获取当前脚本所在目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 将 data_path 设置为当前目录下的 csv 文件夹
            data_path = os.path.join(current_dir, "csv")
            # 将 output_path 设置为当前目录下的 xlsx 文件夹
            output_path = os.path.join(current_dir, "xlsx")
            csv_save_as_xlsx(data_path, output_path) 
            input("\n转换完成，按 Enter 键继续...")
            break
        elif choice == 0:
            return  # Exiting the function, which effectively returns to the main menu
        else:
            print("你的输入有误，请重新输入正确选项！")
            break

def main():
    global local_province_id, local_type_id, school_id, year
    colorama.init(autoreset=True)  # 初始化colorama库
    # 获取用户输入，如果为空则使用默认值
    # 输入有效省市区代码
    print("=========================================")
    print("\n请初始化查询参数\n")
    print(Fore.RED + "(本程序适合重庆、河北、辽宁、江苏、福建、湖北、湖南、广东考生):\n" + Style.RESET_ALL)
    while True:
        local_province_id = input(Fore.GREEN + " ※ 请输入省市区代码" + Fore.RED + "（例如 50，默认值为50）: " + Style.RESET_ALL) or "50"
        if get_province_name(local_province_id) is None:
            print(Fore.RED + "无效的省市区代码，请重新输入。\n" + Style.RESET_ALL)
            #input("按 Enter 键重新输入...\n")
            continue
        else:
            break
    while True:
        year = input(Fore.GREEN + " ※ 请输入年份" + Fore.RED + "（例如 2016 - 2028之间的年份，默认值为2024）: " + Style.RESET_ALL) or "2024"
        if not year.isdigit() or int(year) not in range(2016, 2028):
            print(Fore.RED + "错误：请输入2016 - 2028之间的有效年份。" + Style.RESET_ALL)
            continue
        else:
            break  # 如果输入的年份有效，则退出循环
    if int(year) >= 2021:
        while True:
            # 提示用户输入并获取 local_type_id
            local_type_id = input(Fore.GREEN + " ※ 请输入物理、历史类代码" + Fore.RED + "（2021年及之后，2073 代表物理类，2074 代表历史类，默认值为2073）: " + Style.RESET_ALL) or "2073"
            # 检查 local_type_id 是否在指定的范围内
            if local_type_id in ["2073", "2074"]:
                break  # 如果输入正确，跳出循环
            else:
                print(Fore.RED + "你输入的数字错误，请按照提示重新输入文理科代码！" + Style.RESET_ALL)
                print("2021年之后的文理科代码是：" + Fore.RED + "2073 代表物理类，2074 代表历史类。" + Style.RESET_ALL)
    else:
        while True:
            # 提示用户输入并获取 local_type_id
            local_type_id = input(Fore.GREEN + " ※ 请输入文、理科代码" + Fore.RED + "（2021年之前（不含），1 代表理科，2 代表文科，默认值为1）: " + Style.RESET_ALL) or "1"
            # 检查 local_type_id 是否在指定的范围内
            if local_type_id in ["1", "2"]:
                break  # 如果输入正确，跳出循环
            else:
                print(Fore.RED + "你输入的数字错误，请按照提示重新输入文理科代码！" + Style.RESET_ALL)
                print("2021年之前的文理科代码是：" + Fore.RED + "1 代表理科，2 代表文科；" + Style.RESET_ALL)
    school_id = input(Fore.GREEN + " ※ 请输入学校ID" + Fore.RED + "(默认：东南大学109)" + Style.RESET_ALL + ":") or "109"

    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        print("=========================================")
        print("请输入要查询的选项")
        print(Fore.RED + "(本程序适合重庆、河北、辽宁、江苏、福建、湖北、湖南、广东考生):\n" + Style.RESET_ALL)
        print(Fore.GREEN + " [1] 查询各省分数线")
        print(Fore.GREEN + " [2] 查询专业分数线")
        print(Fore.GREEN + " [3] 查询招生计划")
        print(Fore.GREEN + " [4] 查询开设专业")
        print(Fore.GREEN + " [5] 查询学科评估")
        print(Fore.GREEN + " [6] 一键查询学校全部信息\n")
        print(Fore.GREEN + " [7] 查询省市区代码或学校ID号\n")
        print(Fore.RED + " [8] 重新定义：省市区代码、文理科代码、学校ID、录取年份等参数\n" + Style.RESET_ALL)
        print(Fore.GREEN + " [9] 查询一分一段\n")
        print(Fore.CYAN + " [10] 清空download文件夹")
        print(Fore.CYAN + " [11] 更新学校id(默认不需要执行)\n" + Style.RESET_ALL)
        print(Fore.GREEN + " [12] 将CSV文件批量转换成XLSX文件\n" + Style.RESET_ALL)
        print(Fore.RED + " [0] 退出\n" + Style.RESET_ALL)
        print("=========================================\n")
        try:
            choice = input("请输入有效选项数字:")
            if choice == "":
                print("输入不能为空，请重新输入！")
                input()
                continue
            choice = int(choice)
            if choice == 0:
                break
            else:
                run_code(choice)
        except KeyboardInterrupt:
            print("\n你已中断操作，自动退出本程序！")
            break
        except ValueError:
            print("请输入有效数字！")

if __name__ == "__main__":
    main()
